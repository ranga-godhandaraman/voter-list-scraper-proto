"""Professional Excel workbook generation for reconnaissance outputs."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=14, color="1F4E79")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
ZEBRA = PatternFill("solid", fgColor="D6EAF8")
WARN = PatternFill("solid", fgColor="F9E79F")
OK = PatternFill("solid", fgColor="D5F5E3")
BAD = PatternFill("solid", fgColor="F5B7B1")


def _auto_width(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(10, length + 2), max_width)


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def _excel_safe(value: Any) -> Any:
    """Coerce nested structures to Excel-safe scalars."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    if df.empty:
        ws.cell(start_row, 1, "(no data)")
        return
    # Normalize object columns so openpyxl never sees raw lists/dicts
    safe = df.copy()
    for col in safe.columns:
        safe[col] = safe[col].map(_excel_safe)
    for r_idx, row in enumerate(dataframe_to_rows(safe, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, _excel_safe(value))
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_idx > start_row and (r_idx - start_row) % 2 == 0:
                cell.fill = ZEBRA
    _style_header(ws, start_row)
    ws.freeze_panes = ws.cell(start_row + 1, 1).coordinate
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _kv_sheet(ws, title: str, mapping: dict[str, Any] | list[tuple[str, Any]]) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Key"
    ws["B3"] = "Value"
    _style_header(ws, 3)
    items = mapping.items() if isinstance(mapping, dict) else mapping
    for i, (k, v) in enumerate(items, start=4):
        ws.cell(i, 1, str(k)).border = THIN
        ws.cell(i, 2, str(v)).border = THIN
        if i % 2 == 0:
            ws.cell(i, 1).fill = ZEBRA
            ws.cell(i, 2).fill = ZEBRA
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:B{ws.max_row}"
    _auto_width(ws)


class ExcelReportBuilder:
    """Build the multi-sheet reconnaissance workbook."""

    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path
        self.wb = Workbook()
        # remove default
        default = self.wb.active
        self.wb.remove(default)

    def add_overview(self, overview: dict[str, Any]) -> None:
        ws = self.wb.create_sheet("Overview", 0)
        _kv_sheet(ws, "ECI Download E-Roll — Technical Reconnaissance Overview", overview)

    def add_dataframe(self, name: str, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet(name[:31])
        _write_df(ws, df)

    def add_kv(self, name: str, title: str, data: dict[str, Any]) -> None:
        ws = self.wb.create_sheet(name[:31])
        _kv_sheet(ws, title, data)

    def add_conditional_state_summary(self, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet("State Summary")
        _write_df(ws, df)
        # Highlight zero constituency counts
        if not df.empty and "constituency_count" in df.columns:
            # column index
            col_idx = list(df.columns).index("constituency_count") + 1
            letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                FormulaRule(formula=[f"{letter}2=0"], fill=BAD),
            )
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                FormulaRule(formula=[f"{letter}2>0"], fill=OK),
            )

    def save(self) -> Path:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.output_path)
        logger.info("Excel workbook written to {}", self.output_path)
        return self.output_path


def build_workbook(payload: dict[str, Any], path: Path) -> Path:
    """Assemble all required sheets from recon payload."""
    x = ExcelReportBuilder(path)

    overview = payload.get("overview", {})
    x.add_overview(overview)

    states = pd.DataFrame(payload.get("states", []))
    x.add_conditional_state_summary(states if not states.empty else pd.DataFrame([{"note": "empty"}]))

    districts = pd.DataFrame(payload.get("districts", []))
    if len(districts) > 5000:
        districts = districts.head(5000)
    x.add_dataframe("District Summary", districts if not districts.empty else pd.DataFrame([{"note": "empty"}]))

    acs = pd.DataFrame(payload.get("constituencies", []))
    # Keep workbook usable
    if len(acs) > 8000:
        acs = acs.head(8000)
    x.add_dataframe("Constituencies", acs if not acs.empty else pd.DataFrame([{"note": "empty"}]))

    x.add_dataframe("Revision Details", pd.DataFrame(payload.get("revision_details", [])))
    x.add_dataframe("PDF Availability", pd.DataFrame(payload.get("pdf_availability", [])))
    x.add_dataframe("API Endpoints", pd.DataFrame(payload.get("api_endpoints", [])))
    x.add_dataframe("DOM Elements", pd.DataFrame(payload.get("dom_elements", [])))
    x.add_dataframe("Network Calls", pd.DataFrame(payload.get("network_calls", [])))
    x.add_dataframe("Cookies", pd.DataFrame(payload.get("cookies", [])))
    x.add_dataframe("Headers", pd.DataFrame(payload.get("headers", [])))
    x.add_dataframe("Parameters", pd.DataFrame(payload.get("parameters", [])))
    x.add_kv("Download Flow", "Download Flow Steps", payload.get("download_flow", {}))
    x.add_kv("Website Structure", "Website Structure", payload.get("website_structure", {}))
    x.add_dataframe("Technical Findings", pd.DataFrame(payload.get("technical_findings", [])))
    x.add_kv("Feasibility Report", "Feasibility", payload.get("feasibility_flat", {}))
    x.add_dataframe("Errors", pd.DataFrame(payload.get("errors", [])))
    x.add_dataframe("Recommendations", pd.DataFrame(payload.get("recommendations", [])))
    x.add_dataframe("Risk Analysis", pd.DataFrame(payload.get("risks", [])))

    return x.save()
